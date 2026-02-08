import pandas as pd
from duunitori.pickle_helper import load, pickle_save
from duunitori.class_job import Job
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def style_if_column_false(column_name):
    """Color the row red if its value is false"""

    def _style(row):
        for _ in row:
            return ['background-color: red' if row[column_name] == False else
                    ['background-color: #ffffff' if row.name % 2 else 'background-color: #e6e9fd'][0] for _ in row]
    return _style

def combine_styles(row):
    """Color the row red if its value is false"""
    red_style = style_if_column_false('remote')(row)
    return red_style

def df_to_excel(df:pd.DataFrame, path = "Jobs.xlsx"):
    """To excel wrapper"""
    df.to_excel(path, index=False)
    return path


def move_column_to_last(df, column_name):
    """
    Move a column to the last position in a DataFrame.

    Args:
        df (pd.DataFrame): Input DataFrame
        column_name (str): Column to move
    
    Returns:
        pd.DataFrame: New DataFrame with the column moved
    """
    cols = [col for col in df.columns if col != column_name] + [column_name]
    return df[cols]


def colour_excel(df:pd.DataFrame):
    styled_df = df.style.apply(combine_styles, axis=1)
    return styled_df

def read_excel_sheet(filename, sheet_name=None, values_only=True):
    """
    Read data from an Excel sheet using openpyxl.

    Args:
        filename (str): Path to the Excel file.
        sheet_name (str, optional): Sheet name to read. 
                                    Defaults to the first sheet.
        values_only (bool): If True, return just cell values (no styles).
    
    Returns:
        list of lists: Rows of data from the sheet.
    """
    wb = load_workbook(filename, data_only=True)  # data_only=True -> evaluates formulas
    return wb

def get_column_values(skip_columns):
    col_vals = [
        get_column_letter(c) if isinstance(c, int) else c.upper()
        for c in skip_columns
    ]
    return col_vals

def auto_adjust_column_width(ws, skip_columns=None, padding=2):
    """Auto adjust all columns that are not in the skip columns parameter"""
    if skip_columns is None:
        skip_columns = []
    cols_to_skip = get_column_values(skip_columns)

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter in cols_to_skip:
            continue

        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + padding
    return ws


